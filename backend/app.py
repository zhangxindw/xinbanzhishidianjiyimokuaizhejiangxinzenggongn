import os
import uuid
import random
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
from models import db, Question, Chapter, QuestionType, WrongQuestion, FavoriteQuestion, UserAnswer, UserPreference, OperationLog, KnowledgePoint, KnowledgePointItem, KnowledgePointRelation, MemoryRecord, DistinguishQuestion, DistinguishOption, DistinguishRecord, PracticePlanRecord
from openpyxl import load_workbook, Workbook

# 获取backend目录的父目录（即项目根目录）
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FRONTEND_DIST = os.path.join(BASE_DIR, 'frontend', 'dist')

app = Flask(__name__, static_folder=FRONTEND_DIST, static_url_path='/static')
app.config['SECRET_KEY'] = 'quiz-system-secret-key-2024'
# 数据库使用绝对路径（backend目录下的instance/quiz_system.db），不依赖启动目录
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BACKEND_DIR, 'instance', 'quiz_system.db')
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

CORS(app)
db.init_app(app)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 提供前端静态文件
@app.route('/')
def serve_index():
    return send_from_directory(app.static_folder, 'index.html')

# 全局错误处理器
@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    error_trace = traceback.format_exc()
    print(f"全局异常捕获: {e}")
    print(f"错误堆栈: {error_trace}")
    return jsonify({'status': 'error', 'message': str(e), 'trace': error_trace}), 500

def log_operation(user_id, action, target_type=None, target_id=None, details=None):
    log = OperationLog(
        user_id=user_id,
        action=action,
        target_type=target_type,
        target_id=target_id,
        details=details
    )
    db.session.add(log)
    db.session.commit()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ['xls', 'xlsx']

def to_html(value):
    """将文本转换为HTML，允许安全的HTML标签"""
    if not value:
        return ''
    s = str(value)
    
    import re
    
    # 允许的标签列表（不区分大小写）
    allowed_tags = ['img', 'br', 'p', 'div', 'span', 'strong', 'em', 
                   'b', 'i', 'u', 'ul', 'ol', 'li', 'table', 'tr', 
                   'td', 'th', 'tbody', 'thead', 'h1', 'h2', 'h3', 
                   'h4', 'h5', 'h6', 'a', 'pre', 'code', 'hr', 'sub', 'sup', 'mark']
    
    # 第一步：先处理格式标记（在转义前处理，避免干扰）
    # **加粗** → <strong>加粗</strong>
    s = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', s)
    # __下划线__ → <u>下划线</u>
    s = re.sub(r'__(.+?)__', r'<u>\1</u>', s)
    # ==高亮== → <mark>高亮</mark>
    s = re.sub(r'==(.+?)==', r'<mark style="background-color: #ffff00;">\1</mark>', s)
    # [[关键词]] → <span class="blank-hidden">关键词</span>
    s = re.sub(r'\[\[(.+?)\]\]', r'<span class="blank-hidden">\1</span>', s)
    
    # 第二步：处理换行符
    s = s.replace('\n', '<br>')
    
    # 第三步：标记允许的标签为占位符
    placeholder_map = {}
    placeholder_counter = 0
    tag_regex = re.compile(r'</?(\w+)[^>]*>', re.IGNORECASE)
    
    def tag_replacer(match):
        nonlocal placeholder_counter
        tag = match.group(1).lower()
        if tag in allowed_tags:
            placeholder = f'__SAFE_TAG_{placeholder_counter}__'
            placeholder_map[placeholder] = match.group(0)
            placeholder_counter += 1
            return placeholder
        else:
            # 非允许标签则转义
            return f'&lt;{match.group(0)[1:-1]}&gt;'
    
    s = tag_regex.sub(tag_replacer, s)
    
    # 第四步：转义其他HTML特殊字符
    s = s.replace('&', '&amp;').replace('"', '&quot;').replace("'", '&#39;')
    
    # 第五步：恢复安全标签（将占位符替换回原始标签）
    for placeholder, original_tag in placeholder_map.items():
        s = s.replace(placeholder, original_tag)
    
    return s

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok', 'message': 'Quiz System API is running'})

@app.route('/api/init-db', methods=['POST'])
def init_database():
    db.create_all()
    if QuestionType.query.count() == 0:
        default_types = [
            QuestionType(name='单选题', is_multiple=False),
            QuestionType(name='多选题', is_multiple=True),
            QuestionType(name='判断题', is_multiple=False)
        ]
        db.session.add_all(default_types)
    if Chapter.query.count() == 0:
        default_chapter = Chapter(name='默认章节', order=0)
        db.session.add(default_chapter)
    if UserPreference.query.count() == 0:
        default_pref = UserPreference(user_id='default_user')
        db.session.add(default_pref)
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'Database initialized'})

@app.route('/api/chapters', methods=['GET', 'POST'])
def handle_chapters():
    if request.method == 'GET':
        chapters = Chapter.query.filter_by(parent_id=None).order_by(Chapter.order).all()
        chapters_data = []
        for c in chapters:
            chapter_dict = c.to_dict()
            # 统计该章节的题目数量
            question_count = Question.query.filter_by(chapter_id=c.id).count()
            chapter_dict['question_count'] = question_count
            chapters_data.append(chapter_dict)
        return jsonify({'status': 'ok', 'data': chapters_data})
    else:
        data = request.json
        chapter = Chapter(
            name=data.get('name'),
            parent_id=data.get('parent_id'),
            order=data.get('order', 0)
        )
        db.session.add(chapter)
        db.session.commit()
        log_operation('admin', 'create_chapter', 'chapter', chapter.id, f'Created chapter: {chapter.name}')
        return jsonify({'status': 'ok', 'data': chapter.to_dict()})

@app.route('/api/chapters/<int:chapter_id>', methods=['GET', 'PUT', 'DELETE'])
def handle_chapter(chapter_id):
    chapter = Chapter.query.get_or_404(chapter_id)
    if request.method == 'GET':
        return jsonify({'status': 'ok', 'data': chapter.to_dict()})
    elif request.method == 'PUT':
        data = request.json
        chapter.name = data.get('name', chapter.name)
        chapter.parent_id = data.get('parent_id', chapter.parent_id)
        chapter.order = data.get('order', chapter.order)
        db.session.commit()
        log_operation('admin', 'update_chapter', 'chapter', chapter.id, f'Updated chapter: {chapter.name}')
        return jsonify({'status': 'ok', 'data': chapter.to_dict()})
    else:
        log_operation('admin', 'delete_chapter', 'chapter', chapter.id, f'Deleted chapter: {chapter.name}')
        db.session.delete(chapter)
        db.session.commit()
        return jsonify({'status': 'ok', 'message': 'Chapter deleted'})

@app.route('/api/question-types', methods=['GET', 'POST'])
def handle_question_types():
    if request.method == 'GET':
        types = QuestionType.query.all()
        return jsonify({'status': 'ok', 'data': [t.to_dict() for t in types]})
    else:
        data = request.json
        qtype = QuestionType(
            name=data.get('name'),
            is_multiple=data.get('is_multiple', False)
        )
        db.session.add(qtype)
        db.session.commit()
        log_operation('admin', 'create_question_type', 'question_type', qtype.id, f'Created question type: {qtype.name}')
        return jsonify({'status': 'ok', 'data': qtype.to_dict()})

@app.route('/api/question-types/<int:type_id>', methods=['PUT', 'DELETE'])
def handle_question_type(type_id):
    qtype = QuestionType.query.get_or_404(type_id)
    if request.method == 'PUT':
        data = request.json
        qtype.name = data.get('name', qtype.name)
        qtype.is_multiple = data.get('is_multiple', qtype.is_multiple)
        db.session.commit()
        log_operation('admin', 'update_question_type', 'question_type', qtype.id, f'Updated question type: {qtype.name}')
        return jsonify({'status': 'ok', 'data': qtype.to_dict()})
    else:
        log_operation('admin', 'delete_question_type', 'question_type', qtype.id, f'Deleted question type: {qtype.name}')
        db.session.delete(qtype)
        db.session.commit()
        return jsonify({'status': 'ok', 'message': 'Question type deleted'})

@app.route('/api/questions', methods=['GET', 'POST'])
def handle_questions():
    if request.method == 'GET':
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        stem = request.args.get('stem', '')
        chapter_id = request.args.get('chapter_id', type=int)
        question_type_id = request.args.get('question_type_id', type=int)
        status = request.args.get('status', '')

        query = Question.query
        if stem:
            query = query.filter(Question.stem.contains(stem))
        if chapter_id is not None and chapter_id > 0:
            query = query.filter_by(chapter_id=chapter_id)
        elif chapter_id == 0:
            query = query.filter_by(chapter_id=None)
        if question_type_id:
            query = query.filter_by(question_type_id=question_type_id)
        if status:
            query = query.filter_by(status=status)

        pagination = query.order_by(Question.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
        return jsonify({
            'status': 'ok',
            'data': [q.to_dict() for q in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })
    else:
        data = request.json
        question = Question(
            stem=data.get('stem'),
            option_a=data.get('option_a'),
            option_b=data.get('option_b'),
            option_c=data.get('option_c'),
            option_d=data.get('option_d'),
            option_e=data.get('option_e'),
            option_f=data.get('option_f'),
            answer=data.get('answer', '').upper(),
            explanation=data.get('explanation'),
            stem_html=data.get('stem_html'),
            option_a_html=data.get('option_a_html'),
            option_b_html=data.get('option_b_html'),
            option_c_html=data.get('option_c_html'),
            option_d_html=data.get('option_d_html'),
            option_e_html=data.get('option_e_html'),
            option_f_html=data.get('option_f_html'),
            explanation_html=data.get('explanation_html'),
            difficulty=data.get('difficulty', 1),
            status=data.get('status', 'published'),
            chapter_id=data.get('chapter_id'),
            question_type_id=data.get('question_type_id')
        )
        db.session.add(question)
        db.session.commit()
        log_operation('admin', 'create_question', 'question', question.id, f'Created question: {question.stem[:50]}')
        return jsonify({'status': 'ok', 'data': question.to_dict()})

@app.route('/api/questions/<int:question_id>', methods=['GET', 'PUT', 'DELETE'])
def handle_question(question_id):
    question = Question.query.get_or_404(question_id)
    if request.method == 'GET':
        return jsonify({'status': 'ok', 'data': question.to_dict()})
    elif request.method == 'PUT':
        data = request.json
        question.stem = data.get('stem', question.stem)
        question.option_a = data.get('option_a', question.option_a)
        question.option_b = data.get('option_b', question.option_b)
        question.option_c = data.get('option_c', question.option_c)
        question.option_d = data.get('option_d', question.option_d)
        question.option_e = data.get('option_e', question.option_e)
        question.option_f = data.get('option_f', question.option_f)
        question.answer = data.get('answer', question.answer).upper()
        question.explanation = data.get('explanation', question.explanation)
        question.stem_html = data.get('stem_html', question.stem_html)
        question.option_a_html = data.get('option_a_html', question.option_a_html)
        question.option_b_html = data.get('option_b_html', question.option_b_html)
        question.option_c_html = data.get('option_c_html', question.option_c_html)
        question.option_d_html = data.get('option_d_html', question.option_d_html)
        question.option_e_html = data.get('option_e_html', question.option_e_html)
        question.option_f_html = data.get('option_f_html', question.option_f_html)
        question.explanation_html = data.get('explanation_html', question.explanation_html)
        question.difficulty = data.get('difficulty', question.difficulty)
        question.status = data.get('status', question.status)
        question.chapter_id = data.get('chapter_id', question.chapter_id)
        question.question_type_id = data.get('question_type_id', question.question_type_id)
        db.session.commit()
        log_operation('admin', 'update_question', 'question', question.id, f'Updated question: {question.stem[:50]}')
        return jsonify({'status': 'ok', 'data': question.to_dict()})
    else:
        # 删除关联的错题记录
        WrongQuestion.query.filter_by(question_id=question.id).delete()
        # 删除关联的收藏记录
        FavoriteQuestion.query.filter_by(question_id=question.id).delete()
        # 删除关联的答题记录
        UserAnswer.query.filter_by(question_id=question.id).delete()
        # 记录操作日志
        log_operation('admin', 'delete_question', 'question', question.id, f'Deleted question: {question.stem[:50]}')
        # 删除题目
        db.session.delete(question)
        db.session.commit()
        return jsonify({'status': 'ok', 'message': 'Question deleted'})

@app.route('/api/questions/batch', methods=['POST'])
def batch_operate_questions():
    data = request.json
    question_ids = data.get('question_ids', [])
    action = data.get('action')

    if not question_ids or not action:
        return jsonify({'status': 'error', 'message': 'Missing required parameters'}), 400

    questions = Question.query.filter(Question.id.in_(question_ids)).all()

    if action == 'delete':
        for q in questions:
            # 删除关联的错题记录
            WrongQuestion.query.filter_by(question_id=q.id).delete()
            # 删除关联的收藏记录
            FavoriteQuestion.query.filter_by(question_id=q.id).delete()
            # 删除关联的答题记录
            UserAnswer.query.filter_by(question_id=q.id).delete()
            # 记录操作日志
            log_operation('admin', 'delete_question', 'question', q.id, f'Batch deleted question: {q.stem[:50]}')
            # 删除题目
            db.session.delete(q)
        db.session.commit()
        return jsonify({'status': 'ok', 'message': f'Deleted {len(questions)} questions'})
    elif action == 'move_chapter':
        chapter_id = data.get('chapter_id')
        for q in questions:
            q.chapter_id = chapter_id
        db.session.commit()
        return jsonify({'status': 'ok', 'message': f'Moved {len(questions)} questions to chapter {chapter_id}'})
    elif action == 'change_type':
        question_type_id = data.get('question_type_id')
        for q in questions:
            q.question_type_id = question_type_id
        db.session.commit()
        return jsonify({'status': 'ok', 'message': f'Changed type for {len(questions)} questions'})
    elif action == 'change_status':
        status = data.get('status')
        for q in questions:
            q.status = status
        db.session.commit()
        return jsonify({'status': 'ok', 'message': f'Changed status for {len(questions)} questions'})

    return jsonify({'status': 'error', 'message': 'Unknown action'}), 400

@app.route('/api/import/template', methods=['GET'])
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "题目导入模板"

    headers = ['题干', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '答案', '解析', '题型']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 15

    template_path = os.path.join(app.config['UPLOAD_FOLDER'], 'question_template.xlsx')
    wb.save(template_path)
    return send_file(template_path, as_attachment=True, download_name='question_template.xlsx')

@app.route('/api/import/upload', methods=['POST'])
def upload_questions():
    try:
        print(f"收到上传请求")
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400

        file = request.files['file']
        print(f"文件名: {file.filename}")
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400

        if not allowed_file(file.filename):
            return jsonify({'status': 'error', 'message': 'Invalid file type'}), 400

        duplicate_handling = request.form.get('duplicate_handling', 'skip')
        print(f"重复处理方式: {duplicate_handling}")

        # 为上传的文件生成唯一的临时文件名，避免文件名冲突
        import uuid
        ext = file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        print(f"文件保存成功: {filepath}")

        wb = None
        try:
            # 确保数据库已初始化
            if QuestionType.query.count() == 0:
                default_types = [
                    QuestionType(name='单选题', is_multiple=False),
                    QuestionType(name='多选题', is_multiple=True),
                    QuestionType(name='判断题', is_multiple=False)
                ]
                db.session.add_all(default_types)
                db.session.commit()

            print("正在加载Excel文件...")
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            print(f"表头: {headers}")
            header_map = {}
            for idx, h in enumerate(headers):
                if h:
                    header_map[h.strip()] = idx

            required_fields = ['题干', '选项A', '选项B', '选项C', '选项D', '答案', '题型']
            for field in required_fields:
                if field not in header_map:
                    # 先关闭工作簿再返回错误
                    if wb:
                        wb.close()
                    return jsonify({'status': 'error', 'message': f'Missing required column: {field}'}), 400

            # 处理Excel换行符的函数
            def process_value(value):
                if value is None:
                    return ''
                s = str(value)
                # Excel中的换行符通常是 \n，但有时会是 \r\n
                # 也需要处理单元格内部的换行符
                s = s.replace('\r\n', '\n').replace('\r', '\n')
                return s

            # 将换行符转换为HTML的函数，允许安全的HTML标签
            def to_html(value):
                if not value:
                    return ''
                s = str(value)
                
                import re
                
                # 允许的标签列表（不区分大小写）
                allowed_tags = ['img', 'br', 'p', 'div', 'span', 'strong', 'em', 
                               'b', 'i', 'u', 'ul', 'ol', 'li', 'table', 'tr', 
                               'td', 'th', 'tbody', 'thead', 'h1', 'h2', 'h3', 
                               'h4', 'h5', 'h6', 'a', 'pre', 'code', 'hr', 'sub', 'sup', 'mark']
                
                # 第一步：先处理格式标记（在转义前处理，避免干扰）
                # **加粗** → <strong>加粗</strong>
                s = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', s)
                # __下划线__ → <u>下划线</u>
                s = re.sub(r'__(.+?)__', r'<u>\1</u>', s)
                # ==高亮== → <mark>高亮</mark>
                s = re.sub(r'==(.+?)==', r'<mark style="background-color: #ffff00;">\1</mark>', s)
                # [[关键词]] → <span class="blank-hidden">关键词</span>
                s = re.sub(r'\[\[(.+?)\]\]', r'<span class="blank-hidden">\1</span>', s)
                
                # 第二步：处理换行符
                s = s.replace('\n', '<br>')
                
                # 第三步：标记允许的标签为占位符
                placeholder_map = {}
                placeholder_counter = 0
                tag_regex = re.compile(r'</?(\w+)[^>]*>', re.IGNORECASE)
                
                def tag_replacer(match):
                    nonlocal placeholder_counter
                    tag = match.group(1).lower()
                    if tag in allowed_tags:
                        placeholder = f'__SAFE_TAG_{placeholder_counter}__'
                        placeholder_map[placeholder] = match.group(0)
                        placeholder_counter += 1
                        return placeholder
                    else:
                        # 非允许标签则转义
                        return f'&lt;{match.group(0)[1:-1]}&gt;'
                
                s = tag_regex.sub(tag_replacer, s)
                
                # 第四步：转义其他HTML特殊字符
                s = s.replace('&', '&amp;').replace('"', '&quot;').replace("'", '&#39;')
                
                # 第五步：恢复安全标签（将占位符替换回原始标签）
                for placeholder, original_tag in placeholder_map.items():
                    s = s.replace(placeholder, original_tag)
                
                return s

            success_count = 0
            error_rows = []
            duplicate_count = 0

            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if all(cell.value is None for cell in row):
                    continue

                try:
                    stem = process_value(row[header_map['题干']].value)
                    if not stem or str(stem).strip() == '':
                        error_rows.append({'row': row_num, 'error': '题干为空'})
                        continue

                    option_a = process_value(row[header_map['选项A']].value)
                    option_b = process_value(row[header_map['选项B']].value)
                    option_c = process_value(row[header_map['选项C']].value)
                    option_d = process_value(row[header_map['选项D']].value)
                    option_e = process_value(row[header_map.get('选项E', -1)].value) if '选项E' in header_map else ''
                    option_f = process_value(row[header_map.get('选项F', -1)].value) if '选项F' in header_map else ''
                    answer = str(row[header_map['答案']].value or '').upper()
                    explanation = process_value(row[header_map.get('解析', -1)].value) if '解析' in header_map else ''
                    question_type_name = row[header_map.get('题型', -1)].value if '题型' in header_map else ''

                    answer_clean = ''.join(c for c in answer if c in 'ABCDEF')
                    if not answer_clean:
                        error_rows.append({'row': row_num, 'error': '答案格式错误'})
                        continue

                    chapter_name = row[header_map.get('章节', -1)].value if '章节' in header_map else ''
                    chapter_id = None
                    if chapter_name:
                        chapter = Chapter.query.filter_by(name=chapter_name).first()
                        if chapter:
                            chapter_id = chapter.id

                    existing = Question.query.filter_by(stem=stem).first()
                    if existing:
                        if duplicate_handling == 'skip':
                            duplicate_count += 1
                            continue
                        elif duplicate_handling == '覆盖':
                            existing.option_a = option_a
                            existing.option_b = option_b
                            existing.option_c = option_c
                            existing.option_d = option_d
                            existing.option_e = option_e
                            existing.option_f = option_f
                            existing.answer = answer_clean
                            existing.explanation = explanation
                            # 同时更新HTML版本
                            existing.stem_html = to_html(stem)
                            existing.option_a_html = to_html(option_a)
                            existing.option_b_html = to_html(option_b)
                            existing.option_c_html = to_html(option_c)
                            existing.option_d_html = to_html(option_d)
                            existing.option_e_html = to_html(option_e)
                            existing.option_f_html = to_html(option_f)
                            existing.explanation_html = to_html(explanation)
                            if chapter_id:
                                existing.chapter_id = chapter_id
                            db.session.commit()
                            success_count += 1
                            continue
                        elif duplicate_handling == 'import_all':
                            # import_all模式：即使重复也创建新题目
                            pass

                    qtype = QuestionType.query.filter_by(name=question_type_name).first()
                    if not qtype:
                        qtype = QuestionType.query.first()

                    question = Question(
                        stem=stem,
                        option_a=option_a,
                        option_b=option_b,
                        option_c=option_c,
                        option_d=option_d,
                        option_e=option_e,
                        option_f=option_f,
                        answer=answer_clean,
                        explanation=explanation,
                        # 添加HTML版本，处理换行符
                        stem_html=to_html(stem),
                        option_a_html=to_html(option_a),
                        option_b_html=to_html(option_b),
                        option_c_html=to_html(option_c),
                        option_d_html=to_html(option_d),
                        option_e_html=to_html(option_e),
                        option_f_html=to_html(option_f),
                        explanation_html=to_html(explanation),
                        question_type_id=qtype.id if qtype else None,
                        chapter_id=chapter_id,
                        status='published'
                    )
                    db.session.add(question)
                    success_count += 1

                except Exception as e:
                    error_rows.append({'row': row_num, 'error': str(e)})

            db.session.commit()
            print(f"导入完成: 成功 {success_count} 条, 跳过 {duplicate_count} 条, 错误 {len(error_rows)} 条")
            log_operation('admin', 'import_questions', 'batch', None, f'Imported {success_count} questions, {duplicate_count} duplicates, {len(error_rows)} errors')

            # 先关闭工作簿
            if wb:
                wb.close()

            response = jsonify({
                'status': 'ok',
                'message': f'Import completed',
                'data': {
                    'success_count': success_count,
                    'duplicate_count': duplicate_count,
                    'error_count': len(error_rows),
                    'error_rows': error_rows[:100]
                }
            })
            print("返回响应:", response.get_json())
            return response, 200

        except Exception as e:
            db.session.rollback()
            print(f"导入处理异常: {str(e)}")
            import traceback
            traceback.print_exc()
            # 确保关闭工作簿
            if wb:
                try:
                    wb.close()
                except:
                    pass
            return jsonify({'status': 'error', 'message': f'Import failed: {str(e)}'}), 500
        finally:
            # 延迟删除文件，避免被占用
            if os.path.exists(filepath):
                print(f"准备删除临时文件: {filepath}")
                try:
                    import time
                    time.sleep(0.1)  # 短暂延迟
                    os.remove(filepath)
                    print(f"临时文件已删除")
                except Exception as e:
                    print(f"删除临时文件失败: {str(e)}")
    except Exception as e:
        print(f"服务器异常: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'Server error: {str(e)}'}), 500

@app.route('/api/export/questions', methods=['POST'])
def export_questions():
    data = request.json
    question_ids = data.get('question_ids', [])

    if question_ids:
        questions = Question.query.filter(Question.id.in_(question_ids)).all()
    else:
        questions = Question.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "题目导出"

    headers = ['题干', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '答案', '解析', '题型', '章节', '难度', '状态']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_num, q in enumerate(questions, 2):
        ws.cell(row=row_num, column=1, value=q.stem)
        ws.cell(row=row_num, column=2, value=q.option_a)
        ws.cell(row=row_num, column=3, value=q.option_b)
        ws.cell(row=row_num, column=4, value=q.option_c)
        ws.cell(row=row_num, column=5, value=q.option_d)
        ws.cell(row=row_num, column=6, value=q.option_e)
        ws.cell(row=row_num, column=7, value=q.option_f)
        ws.cell(row=row_num, column=8, value=q.answer)
        ws.cell(row=row_num, column=9, value=q.explanation)
        ws.cell(row=row_num, column=10, value=q.question_type.name if q.question_type else '')
        ws.cell(row=row_num, column=11, value=q.chapter.name if q.chapter else '')
        ws.cell(row=row_num, column=12, value=q.difficulty)
        ws.cell(row=row_num, column=13, value=q.status)

    export_path = os.path.join(app.config['UPLOAD_FOLDER'], 'exported_questions.xlsx')
    wb.save(export_path)
    return send_file(export_path, as_attachment=True, download_name='exported_questions.xlsx')

@app.route('/api/practice/memorize', methods=['GET'])
def practice_memorize():
    chapter_ids = request.args.getlist('chapter_ids', type=int)
    chapter_id = request.args.get('chapter_id', type=int)
    question_type_id = request.args.get('question_type_id', type=int)

    query = Question.query.filter_by(status='published')
    if chapter_ids:
        query = query.filter(Question.chapter_id.in_(chapter_ids))
    elif chapter_id:
        query = query.filter_by(chapter_id=chapter_id)
    if question_type_id:
        query = query.filter_by(question_type_id=question_type_id)

    questions = query.order_by(Question.id).all()
    return jsonify({
        'status': 'ok',
        'data': [q.to_dict() for q in questions],
        'total': len(questions)
    })

@app.route('/api/practice/sequential', methods=['POST'])
def practice_sequential():
    data = request.json
    chapter_ids = data.get('chapter_ids', [])
    question_type_id = data.get('question_type_id')
    shuffle_options = data.get('shuffle_options', False)

    query = Question.query.filter_by(status='published')
    if chapter_ids:
        query = query.filter(Question.chapter_id.in_(chapter_ids))
    if question_type_id:
        query = query.filter_by(question_type_id=question_type_id)

    questions = query.order_by(Question.id).all()
    result = []
    for q in questions:
        q_dict = q.to_dict(shuffle_options=shuffle_options)
        result.append(q_dict)

    session_id = str(uuid.uuid4())
    return jsonify({
        'status': 'ok',
        'data': result,
        'session_id': session_id
    })

@app.route('/api/practice/random', methods=['POST'])
def practice_random():
    data = request.json
    count = data.get('count', 10)
    score = data.get('score', 5)
    shuffle_options = data.get('shuffle_options', True)
    chapter_ids = data.get('chapter_ids', [])
    chapter_ratios = data.get('chapter_ratios', {})
    wrong_ratio = data.get('wrong_ratio', 0)
    user_id = data.get('user_id', 'default_user')
    question_type_id = data.get('question_type_id')

    selected = []
    selected_ids = set()

    # 获取错题本中的易错题（错误次数>2）
    wrong_question_ids = []
    if wrong_ratio > 0:
        wrong_query = db.session.query(
            WrongQuestion.question_id,
            db.func.count(WrongQuestion.id).label('wrong_count')
        ).filter(
            WrongQuestion.user_id == user_id
        ).group_by(
            WrongQuestion.question_id
        ).having(
            db.func.count(WrongQuestion.id) > 2
        ).subquery()
        wrong_question_ids = [wq.question_id for wq in db.session.query(wrong_query).all()]

    # 判断是否设置了章节抽题比例
    has_chapter_ratios = any(chapter_ratios.get(cid, 0) > 0 for cid in chapter_ids) if chapter_ids else False

    # 计算各章节的抽题数量
    chapter_totals = {}
    if chapter_ids:
        if has_chapter_ratios:
            # 情况一：部分或全部章节设置了比例
            final_chapter_ratios = {}
            set_chapters = [cid for cid in chapter_ids if chapter_ratios.get(cid, 0) > 0]
            unset_chapters = [cid for cid in chapter_ids if chapter_ratios.get(cid, 0) == 0]

            # 设置过比例的章节
            for cid in set_chapters:
                final_chapter_ratios[cid] = chapter_ratios[cid]

            # 未设置比例的章节平均分配剩余比例
            if unset_chapters:
                remaining = 100 - sum(final_chapter_ratios.values())
                if remaining > 0:
                    each_ratio = int(remaining / len(unset_chapters))
                    for i, cid in enumerate(unset_chapters):
                        if i == len(unset_chapters) - 1:
                            final_chapter_ratios[cid] = remaining - each_ratio * (len(unset_chapters) - 1)
                        else:
                            final_chapter_ratios[cid] = each_ratio

            # 计算各章节抽题数量
            for cid in chapter_ids:
                chapter_totals[cid] = int(count * final_chapter_ratios.get(cid, 0) / 100)
        else:
            # 情况二：都没设置比例，不分配具体章节数量，最后统一从选中章节随机抽取
            pass
    else:
        # 没有选中章节，从全部题库抽取
        pass

    # 获取所有可选题目的基础查询
    base_query = Question.query.filter_by(status='published')
    if question_type_id:
        base_query = base_query.filter_by(question_type_id=question_type_id)

    # 开始抽题
    if chapter_ids:
        if has_chapter_ratios:
            # 按章节比例抽题
            for chapter_id in chapter_ids:
                chapter_total = chapter_totals.get(chapter_id, 0)
                if chapter_total <= 0:
                    continue

                # 获取该章节的题目
                chapter_questions = base_query.filter_by(chapter_id=chapter_id).all()
                if not chapter_questions:
                    continue

                # 从该章节抽取题目
                chapter_selected = []

                # 1. 先从该章节抽易错题（如果设置了易错题比例）
                if wrong_ratio > 0:
                    chapter_wrong_count = int(chapter_total * wrong_ratio / 100)
                    if chapter_wrong_count > 0:
                        chapter_wrong_questions = [
                            q for q in chapter_questions
                            if q.id in wrong_question_ids and q.id not in selected_ids
                        ]
                        if chapter_wrong_questions:
                            take = min(chapter_wrong_count, len(chapter_wrong_questions))
                            chapter_selected.extend(random.sample(chapter_wrong_questions, take))

                # 2. 用该章节普通题目补足
                remaining = chapter_total - len(chapter_selected)
                if remaining > 0:
                    chapter_normal_questions = [
                        q for q in chapter_questions
                        if q.id not in selected_ids and q not in chapter_selected
                    ]
                    if chapter_normal_questions:
                        take = min(remaining, len(chapter_normal_questions))
                        chapter_selected.extend(random.sample(chapter_normal_questions, take))

                selected.extend(chapter_selected)
                selected_ids.update(q.id for q in chapter_selected)
        else:
            # 没有设置章节比例，统一从选中章节抽题
            # 获取所有选中章节的题目
            all_chapter_questions = base_query.filter(Question.chapter_id.in_(chapter_ids)).all()
            if not all_chapter_questions:
                pass
            else:
                # 1. 先抽易错题
                if wrong_ratio > 0:
                    wrong_count = int(count * wrong_ratio / 100)
                    chapter_wrong_questions = [
                        q for q in all_chapter_questions
                        if q.id in wrong_question_ids and q.id not in selected_ids
                    ]
                    if chapter_wrong_questions:
                        take = min(wrong_count, len(chapter_wrong_questions))
                        selected.extend(random.sample(chapter_wrong_questions, take))
                        selected_ids.update(q.id for q in selected)

                # 2. 用普通题目补足
                remaining = count - len(selected)
                if remaining > 0:
                    available_normal = [
                        q for q in all_chapter_questions
                        if q.id not in selected_ids
                    ]
                    if available_normal:
                        take = min(remaining, len(available_normal))
                        selected.extend(random.sample(available_normal, take))

        # 如果从章节抽取后总数量不足，从选中章节中继续补足
        if len(selected) < count:
            remaining = count - len(selected)
            all_chapter_questions = base_query.filter(Question.chapter_id.in_(chapter_ids)).all()
            remaining_questions = [q for q in all_chapter_questions if q.id not in selected_ids]

            if remaining_questions:
                take = min(remaining, len(remaining_questions))
                selected.extend(random.sample(remaining_questions, take))
    else:
        # 没有选中章节，从全部题库抽取
        all_questions = base_query.all()

        # 1. 先抽易错题
        if wrong_ratio > 0:
            wrong_count = int(count * wrong_ratio / 100)
            available_wrong = [q for q in all_questions if q.id in wrong_question_ids and q.id not in selected_ids]
            if available_wrong:
                take = min(wrong_count, len(available_wrong))
                selected.extend(random.sample(available_wrong, take))
                selected_ids.update(q.id for q in selected)

        # 2. 用普通题目补足
        remaining = count - len(selected)
        if remaining > 0:
            available_normal = [q for q in all_questions if q.id not in selected_ids]
            if available_normal:
                take = min(remaining, len(available_normal))
                selected.extend(random.sample(available_normal, take))

    # 打乱最终结果
    random.shuffle(selected)

    # 构建返回结果
    result = []
    for q in selected[:count]:
        q_dict = q.to_dict(shuffle_options=shuffle_options)
        q_dict['score'] = score
        result.append(q_dict)

    session_id = str(uuid.uuid4())
    return jsonify({
        'status': 'ok',
        'data': result,
        'session_id': session_id,
        'total': len(result)
    })

@app.route('/api/wrong-questions', methods=['GET', 'POST'])
def handle_wrong_questions():
    user_id = request.args.get('user_id', 'default_user')

    if request.method == 'GET':
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        question_type_id = request.args.get('question_type_id', type=int)
        chapter_id = request.args.get('chapter_id', type=int)
        wrong_count_eq = request.args.get('wrong_count_eq', type=int)
        min_reappearance_count = request.args.get('min_reappearance_count', type=int)

        query = WrongQuestion.query.filter_by(user_id=user_id)
        if question_type_id:
            query = query.join(Question).filter(Question.question_type_id == question_type_id)
        if chapter_id:
            query = query.join(Question).filter(Question.chapter_id == chapter_id)
        if wrong_count_eq is not None:
            if wrong_count_eq == 1:
                query = query.filter(WrongQuestion.wrong_count == 1)
            elif wrong_count_eq == 2:
                query = query.filter(WrongQuestion.wrong_count == 2)
            elif wrong_count_eq == 3:
                query = query.filter(WrongQuestion.wrong_count == 3)
            elif wrong_count_eq == 4:
                query = query.filter(WrongQuestion.wrong_count > 3)
            elif wrong_count_eq == 5:
                query = query.filter(WrongQuestion.wrong_count == 4)
            elif wrong_count_eq == 6:
                query = query.filter(WrongQuestion.wrong_count == 5)
            elif wrong_count_eq == 7:
                query = query.filter(WrongQuestion.wrong_count > 5)
            elif wrong_count_eq == 8:
                query = query.filter(WrongQuestion.wrong_count >= 2)
            elif wrong_count_eq == 9:
                query = query.filter(WrongQuestion.wrong_count >= 3)
            elif wrong_count_eq == 10:
                query = query.filter(WrongQuestion.wrong_count >= 4)
            else:
                query = query.filter(WrongQuestion.wrong_count == wrong_count_eq)
        if min_reappearance_count is not None:
            query = query.filter(WrongQuestion.reappearance_count >= min_reappearance_count)

        pagination = query.order_by(WrongQuestion.last_wrong_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
        return jsonify({
            'status': 'ok',
            'data': [wq.to_dict() for wq in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })
    else:
        data = request.json
        question_id = data.get('question_id')
        wrong_q = WrongQuestion(
            user_id=user_id,
            question_id=question_id
        )
        db.session.add(wrong_q)
        db.session.commit()
        return jsonify({'status': 'ok', 'data': wrong_q.to_dict()})

@app.route('/api/wrong-questions/<int:wrong_id>', methods=['DELETE'])
def delete_wrong_question(wrong_id):
    wrong_q = WrongQuestion.query.get_or_404(wrong_id)
    db.session.delete(wrong_q)
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'Removed from wrong questions'})

@app.route('/api/wrong-questions/clear', methods=['DELETE'])
def clear_wrong_questions():
    user_id = request.args.get('user_id', 'default_user')
    WrongQuestion.query.filter_by(user_id=user_id).delete()
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'Cleared all wrong questions'})

@app.route('/api/wrong-questions/batch', methods=['DELETE'])
def batch_remove_wrong_questions():
    user_id = request.json.get('user_id', 'default_user')
    wrong_ids = request.json.get('wrong_ids', [])
    
    if not wrong_ids:
        return jsonify({'status': 'error', 'message': 'No wrong question IDs provided'}), 400
    
    deleted_count = WrongQuestion.query.filter(
        WrongQuestion.user_id == user_id,
        WrongQuestion.id.in_(wrong_ids)
    ).delete(synchronize_session=False)
    
    db.session.commit()
    return jsonify({
        'status': 'ok', 
        'message': f'Removed {deleted_count} questions from wrong questions book'
    })

@app.route('/api/wrong-questions/stats', methods=['GET'])
def wrong_questions_stats():
    user_id = request.args.get('user_id', 'default_user')
    
    # 获取总错题数
    total = WrongQuestion.query.filter_by(user_id=user_id).count()
    
    # 获取各章节错题统计
    chapter_stats = db.session.query(
        Question.chapter_id,
        db.func.count(WrongQuestion.id).label('count')
    ).join(
        Question, WrongQuestion.question_id == Question.id
    ).filter(
        WrongQuestion.user_id == user_id,
        Question.chapter_id.isnot(None)
    ).group_by(
        Question.chapter_id
    ).all()
    
    chapter_stats_dict = {stat.chapter_id: stat.count for stat in chapter_stats}
    
    return jsonify({
        'status': 'ok',
        'data': {
            'total': total,
            'chapter_stats': chapter_stats_dict
        }
    })

@app.route('/api/wrong-questions/practice', methods=['POST'])
def practice_wrong_questions():
    user_id = request.json.get('user_id', 'default_user')
    shuffle = request.json.get('shuffle', True)
    shuffle_options = request.json.get('shuffle_options', True)
    chapter_ids = request.json.get('chapter_ids', [])
    wrong_ids = request.json.get('wrong_ids', [])
    increment_reappearance = request.json.get('increment_reappearance', True)

    query = WrongQuestion.query.filter_by(user_id=user_id)
    
    if chapter_ids and len(chapter_ids) > 0:
        query = query.join(Question).filter(Question.chapter_id.in_(chapter_ids))
    
    if wrong_ids and len(wrong_ids) > 0:
        query = query.filter(WrongQuestion.id.in_(wrong_ids))
    
    wrong_questions = query.all()
    
    if not wrong_questions:
        return jsonify({'status': 'ok', 'data': [], 'session_id': str(uuid.uuid4())})

    if increment_reappearance:
        for wq in wrong_questions:
            wq.reappearance_count += 1
        db.session.commit()

    questions = [wq.question for wq in wrong_questions if wq.question]
    if shuffle:
        random.shuffle(questions)

    result = []
    for q in questions:
        q_dict = q.to_dict(shuffle_options=shuffle_options)
        result.append(q_dict)

    session_id = str(uuid.uuid4())
    return jsonify({
        'status': 'ok',
        'data': result,
        'session_id': session_id
    })

@app.route('/api/favorites', methods=['GET', 'POST'])
def handle_favorites():
    user_id = request.args.get('user_id', 'default_user')

    if request.method == 'GET':
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)

        pagination = FavoriteQuestion.query.filter_by(user_id=user_id).order_by(
            FavoriteQuestion.created_at.desc()
        ).paginate(page=page, per_page=per_page, error_out=False)

        return jsonify({
            'status': 'ok',
            'data': [f.to_dict() for f in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })
    else:
        data = request.json
        question_id = data.get('question_id')

        existing = FavoriteQuestion.query.filter_by(user_id=user_id, question_id=question_id).first()
        if existing:
            return jsonify({'status': 'ok', 'data': existing.to_dict(), 'message': 'Already favorited'})

        favorite = FavoriteQuestion(
            user_id=user_id,
            question_id=question_id
        )
        db.session.add(favorite)
        db.session.commit()
        return jsonify({'status': 'ok', 'data': favorite.to_dict()})

@app.route('/api/favorites/<int:favorite_id>', methods=['DELETE'])
def delete_favorite(favorite_id):
    favorite = FavoriteQuestion.query.get_or_404(favorite_id)
    db.session.delete(favorite)
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'Removed from favorites'})

@app.route('/api/favorites/check/<int:question_id>', methods=['GET'])
def check_favorite(question_id):
    user_id = request.args.get('user_id', 'default_user')
    favorite = FavoriteQuestion.query.filter_by(user_id=user_id, question_id=question_id).first()
    return jsonify({'status': 'ok', 'is_favorite': favorite is not None})

@app.route('/api/answers', methods=['POST'])
def submit_answer():
    data = request.json
    session_id = data.get('session_id', 'default')
    question_id = data.get('question_id')
    user_answer = data.get('answer', '').upper()
    user_id = data.get('user_id', 'default_user')
    expected_answer = data.get('expected_answer')

    question = Question.query.get_or_404(question_id)
    if expected_answer is not None and expected_answer != '':
        correct_answer = expected_answer.upper()
    else:
        correct_answer = question.answer.upper()
    is_correct = user_answer == correct_answer

    answer = UserAnswer(
        session_id=session_id,
        question_id=question_id,
        user_answer=user_answer,
        is_correct=is_correct
    )
    db.session.add(answer)

    if not is_correct:
        wrong_q = WrongQuestion.query.filter_by(user_id=user_id, question_id=question_id).first()
        if wrong_q:
            wrong_q.wrong_count += 1
            wrong_q.last_wrong_at = datetime.utcnow()
        else:
            wrong_q = WrongQuestion(
                user_id=user_id,
                question_id=question_id
            )
            db.session.add(wrong_q)

    db.session.commit()

    return jsonify({
        'status': 'ok',
        'data': {
            'is_correct': is_correct,
            'correct_answer': correct_answer,
            'explanation': question.explanation,
            'explanation_html': question.explanation_html
        }
    })

@app.route('/api/answers/history/<int:question_id>', methods=['GET'])
def get_answer_history(question_id):
    user_id = request.args.get('user_id', 'default_user')
    answers = UserAnswer.query.filter_by(session_id=user_id, question_id=question_id).order_by(
        UserAnswer.answered_at.desc()
    ).limit(10).all()
    return jsonify({
        'status': 'ok',
        'data': [a.to_dict() for a in answers]
    })

@app.route('/api/preferences', methods=['GET', 'PUT'])
def handle_preferences():
    user_id = request.args.get('user_id', 'default_user')

    pref = UserPreference.query.filter_by(user_id=user_id).first()
    if not pref:
        pref = UserPreference(user_id=user_id)
        db.session.add(pref)
        db.session.commit()

    if request.method == 'GET':
        return jsonify({'status': 'ok', 'data': pref.to_dict()})
    else:
        data = request.json
        pref.eye_protection_mode = data.get('eye_protection_mode', pref.eye_protection_mode)
        pref.font_size = data.get('font_size', pref.font_size)
        pref.font_family = data.get('font_family', pref.font_family)
        db.session.commit()
        return jsonify({'status': 'ok', 'data': pref.to_dict()})

@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    user_id = request.args.get('user_id', 'default_user')

    total_questions = Question.query.filter_by(status='published').count()
    total_answered = UserAnswer.query.filter_by(session_id=user_id).count()
    total_correct = UserAnswer.query.filter_by(session_id=user_id, is_correct=True).count()
    correct_rate = (total_correct / total_answered * 100) if total_answered > 0 else 0

    wrong_count = WrongQuestion.query.filter_by(user_id=user_id).count()
    favorite_count = FavoriteQuestion.query.filter_by(user_id=user_id).count()

    type_stats = db.session.query(
        Question.question_type_id,
        QuestionType.name,
        db.func.count(Question.id).label('count')
    ).join(QuestionType).group_by(Question.question_type_id).all()

    return jsonify({
        'status': 'ok',
        'data': {
            'total_questions': total_questions,
            'total_answered': total_answered,
            'total_correct': total_correct,
            'correct_rate': round(correct_rate, 2),
            'wrong_count': wrong_count,
            'favorite_count': favorite_count,
            'type_stats': [{'type_id': t[0], 'type_name': t[1], 'count': t[2]} for t in type_stats]
        }
    })

@app.route('/api/logs', methods=['GET'])
def get_operation_logs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    pagination = OperationLog.query.order_by(OperationLog.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        'status': 'ok',
        'data': [log.to_dict() for log in pagination.items],
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages
    })

# ========== 知识点记忆模块 API ==========

# 知识点列表
@app.route('/api/knowledge-points', methods=['GET', 'POST'])
def handle_knowledge_points():
    if request.method == 'GET':
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        chapter_id = request.args.get('chapter_id', type=int)
        chapter_ids = request.args.getlist('chapter_ids', type=int)  # 支持多章节筛选
        priority = request.args.get('priority')
        keyword = request.args.get('keyword')
        user_id = request.args.get('user_id', 'default_user')
        exclude_in_plan = request.args.get('exclude_in_plan', 'false').lower() == 'true'

        # 兼容旧数据：查询 status 为 'active' 或 'published' 的知识点
        query = KnowledgePoint.query.filter(KnowledgePoint.status.in_(['active', 'published']))
        
        if chapter_id:
            query = query.filter_by(chapter_id=chapter_id)
        elif chapter_ids:  # 支持多章节筛选
            query = query.filter(KnowledgePoint.chapter_id.in_(chapter_ids))
        if priority:
            query = query.filter_by(priority=priority)
        if keyword:
            query = query.filter(KnowledgePoint.question.like('%' + keyword + '%'))

        # 如果需要排除已加入规划的知识点，先获取已加入规划的知识点ID
        excluded_ids = set()
        if exclude_in_plan:
            memory_records = MemoryRecord.query.filter_by(user_id=user_id).all()
            excluded_ids = {mr.knowledge_point_id for mr in memory_records}
            query = query.filter(~KnowledgePoint.id.in_(excluded_ids))

        pagination = query.order_by(KnowledgePoint.chapter_id, KnowledgePoint.created_at).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # 获取用户的记忆记录（用于标记 in_plan 状态）
        memory_records = {mr.knowledge_point_id: mr for mr in MemoryRecord.query.filter_by(user_id=user_id).all()}
        
        result = []
        for kp in pagination.items:
            kp_dict = kp.to_dict()
            # 添加是否加入规划和下一次复习时间
            if kp.id in memory_records:
                mr = memory_records[kp.id]
                kp_dict['in_plan'] = True
                kp_dict['next_review'] = mr.next_review_date.isoformat() if mr.next_review_date else None
            else:
                kp_dict['in_plan'] = False
                kp_dict['next_review'] = None
            result.append(kp_dict)
        
        return jsonify({
            'status': 'ok',
            'data': result,
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })
    else:
        try:
            print(f"=== 收到 POST 请求 ===")
            print(f"请求方法: {request.method}")
            print(f"Content-Type: {request.headers.get('Content-Type')}")
            print(f"request.json: {request.json}")
            print(f"request.data: {request.data}")
            
            if request.json is None:
                print("ERROR: request.json is None!")
                return jsonify({'status': 'error', 'message': '无法解析请求体'}), 400
                
            data = request.json
            question = data.get('question', '')
            priority = data.get('priority', 'normal')
            mnemonic = data.get('mnemonic', '')
            chapter_id = data.get('chapter_id')
            items = data.get('items', [])
            print(f"解析参数: question={question}, priority={priority}, chapter_id={chapter_id}, items={items}")

            # 处理 chapter_id：如果是空字符串，设为 None
            if chapter_id == '':
                chapter_id = None

            kp = KnowledgePoint(
                question=question,
                question_html=to_html(question),
                priority=priority,
                mnemonic=mnemonic,
                chapter_id=chapter_id
            )
            db.session.add(kp)
            db.session.flush()

            for idx, item in enumerate(items):
                kp_item = KnowledgePointItem(
                    knowledge_point_id=kp.id,
                    content=item.get('content', ''),
                    content_html=to_html(item.get('content', '')),
                    order=idx,
                    blank_positions=item.get('blank_positions')
                )
                db.session.add(kp_item)
                db.session.flush()  # 获取 kp_item.id
                
                # 处理关联
                relations = item.get('relations', [])
                for target_id in relations:
                    if target_id != kp.id:  # 不能关联自己
                        relation = KnowledgePointRelation(
                            source_item_id=kp_item.id,
                            target_kp_id=target_id
                        )
                        db.session.add(relation)

            db.session.commit()
            return jsonify({'status': 'ok', 'data': kp.to_dict()})
        except Exception as e:
            db.session.rollback()
            import traceback
            error_trace = traceback.format_exc()
            print(f"创建知识点失败: {str(e)}")
            print(f"错误堆栈: {error_trace}")
            return jsonify({'status': 'error', 'message': str(e), 'trace': error_trace}), 500

# 知识点详情
@app.route('/api/knowledge-points/<int:kp_id>', methods=['GET', 'PUT', 'DELETE'])
def handle_knowledge_point(kp_id):
    kp = KnowledgePoint.query.get_or_404(kp_id)
    
    if request.method == 'GET':
        return jsonify({'status': 'ok', 'data': kp.to_dict()})
    elif request.method == 'PUT':
        data = request.json
        kp.question = data.get('question', kp.question)
        # 如果前端发送了 question_html，直接使用；否则从 question 计算
        if 'question_html' in data:
            kp.question_html = data['question_html']
        else:
            kp.question_html = to_html(kp.question)
        kp.priority = data.get('priority', kp.priority)
        kp.mnemonic = data.get('mnemonic', kp.mnemonic)
        chapter_id = data.get('chapter_id', kp.chapter_id)
        if chapter_id == '':
            chapter_id = None
        kp.chapter_id = chapter_id
        
        # 更新条目（只有当items数组有内容时才处理）
        items = data.get('items', [])
        updated_ids = set()
        
        if items:  # 只有当items有内容时才处理
            existing_items = {item.id: item for item in kp.items.all()}
            
            for idx, item in enumerate(items):
                if item.get('id') and item['id'] in existing_items:
                    # 更新现有条目
                    existing = existing_items[item['id']]
                    new_content = item.get('content', existing.content)
                    existing.content = new_content
                    # 如果前端发送了 content_html，直接使用；否则从新的 content 计算
                    if 'content_html' in item:
                        existing.content_html = item['content_html']
                    else:
                        existing.content_html = to_html(new_content)
                    existing.order = idx
                    existing.blank_positions = item.get('blank_positions', existing.blank_positions)
                    updated_ids.add(item['id'])
                    
                    # 更新关联：先删除旧的关联
                    KnowledgePointRelation.query.filter_by(source_item_id=existing.id).delete()
                    # 添加新的关联
                    relations = item.get('relations', [])
                    for target_id in relations:
                        if target_id != kp.id:
                            relation = KnowledgePointRelation(
                                source_item_id=existing.id,
                                target_kp_id=target_id
                            )
                            db.session.add(relation)
                else:
                    # 创建新条目
                    kp_item = KnowledgePointItem(
                        knowledge_point_id=kp.id,
                        content=item.get('content', ''),
                        content_html=to_html(item.get('content', '')),
                        order=idx,
                        blank_positions=item.get('blank_positions')
                    )
                    db.session.add(kp_item)
                    db.session.flush()
                    
                    # 添加关联
                    relations = item.get('relations', [])
                    for target_id in relations:
                        if target_id != kp.id:
                            relation = KnowledgePointRelation(
                                source_item_id=kp_item.id,
                                target_kp_id=target_id
                            )
                            db.session.add(relation)
            
            # 删除不在新列表中的旧条目
            for item_id, item in existing_items.items():
                if item_id not in updated_ids:
                    db.session.delete(item)
        
        db.session.commit()
        return jsonify({'status': 'ok', 'data': kp.to_dict()})
    else:
        # 先删除所有关联关系
        KnowledgePointRelation.query.filter_by(source_id=kp_id).delete()
        KnowledgePointRelation.query.filter_by(target_id=kp_id).delete()
        # 删除记忆记录
        MemoryRecord.query.filter_by(knowledge_point_id=kp_id).delete()
        # 删除知识点
        db.session.delete(kp)
        db.session.commit()
        return jsonify({'status': 'ok', 'message': 'Knowledge point deleted'})

# 批量删除知识点
@app.route('/api/knowledge-points/batch', methods=['DELETE'])
def batch_delete_knowledge_points():
    """批量删除知识点"""
    data = request.json
    ids = data.get('ids', [])
    
    if not ids:
        return jsonify({'status': 'error', 'message': 'No IDs provided'}), 400
    
    try:
        # 批量删除知识点及其关联数据
        for kp_id in ids:
            kp = KnowledgePoint.query.get(kp_id)
            if kp:
                # 删除所有关联关系
                KnowledgePointRelation.query.filter_by(source_id=kp_id).delete()
                KnowledgePointRelation.query.filter_by(target_id=kp_id).delete()
                # 删除记忆记录
                MemoryRecord.query.filter_by(knowledge_point_id=kp_id).delete()
                # 删除知识点（会级联删除条目）
                db.session.delete(kp)
        
        db.session.commit()
        return jsonify({'status': 'ok', 'message': f'成功删除 {len(ids)} 个知识点'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# 获取条目的关联信息
@app.route('/api/knowledge-point-items/<int:item_id>/relations', methods=['GET'])
def get_item_relations(item_id):
    relations = KnowledgePointRelation.query.filter_by(source_item_id=item_id).all()
    result = []
    for rel in relations:
        target_kp = KnowledgePoint.query.get(rel.target_kp_id)
        if target_kp:
            result.append({
                'id': target_kp.id,
                'question': target_kp.question
            })
    return jsonify({'status': 'ok', 'data': result})

# 更新单个知识点条目
@app.route('/api/knowledge-points/<int:kp_id>/items/<int:item_index>', methods=['PUT'])
def update_knowledge_point_item(kp_id, item_index):
    kp = KnowledgePoint.query.get_or_404(kp_id)
    items = KnowledgePointItem.query.filter_by(knowledge_point_id=kp_id).order_by(KnowledgePointItem.order).all()
    
    if item_index < 0 or item_index >= len(items):
        return jsonify({'status': 'error', 'message': 'Item index out of range'}), 400
    
    item = items[item_index]
    data = request.json
    
    try:
        if 'content' in data:
            item.content = data['content']
            # 如果前端发送了 content_html，直接使用；否则从 content 计算
            if 'content_html' in data:
                item.content_html = data['content_html']
            else:
                item.content_html = to_html(item.content)
        
        db.session.commit()
        return jsonify({'status': 'ok', 'data': item.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# 获取知识点关联关系
@app.route('/api/knowledge-points/<int:kp_id>/relations', methods=['GET', 'POST'])
def handle_kp_relations(kp_id):
    if request.method == 'GET':
        relations = KnowledgePointRelation.query.filter_by(source_id=kp_id).all()
        return jsonify({'status': 'ok', 'data': [r.to_dict() for r in relations]})
    else:
        data = request.json
        target_id = data.get('target_id')
        
        # 检查是否已存在关联
        existing = KnowledgePointRelation.query.filter_by(source_id=kp_id, target_id=target_id).first()
        if existing:
            return jsonify({'status': 'error', 'message': 'Relation already exists'}), 400
        
        relation = KnowledgePointRelation(source_id=kp_id, target_id=target_id)
        db.session.add(relation)
        db.session.commit()
        return jsonify({'status': 'ok', 'data': relation.to_dict()})

# 删除知识点关联
@app.route('/api/knowledge-point-relations/<int:relation_id>', methods=['DELETE'])
def delete_kp_relation(relation_id):
    relation = KnowledgePointRelation.query.get_or_404(relation_id)
    db.session.delete(relation)
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'Relation deleted'})

# 获取知识点被关联的信息（反向查询：哪些答案条目关联了这个知识点）
@app.route('/api/knowledge-points/<int:kp_id>/referenced-by', methods=['GET'])
def get_kp_referenced_by(kp_id):
    """获取知识点被哪些答案条目关联"""
    # 查询所有 target_kp_id 或 target_id 等于 kp_id 的关联记录
    relations = KnowledgePointRelation.query.filter(
        (KnowledgePointRelation.target_kp_id == kp_id) | 
        (KnowledgePointRelation.target_id == kp_id)
    ).all()
    
    result = []
    for rel in relations:
        # 获取源条目信息
        if rel.source_item_id:
            source_item = KnowledgePointItem.query.get(rel.source_item_id)
            if source_item:
                source_kp = KnowledgePoint.query.get(source_item.knowledge_point_id)
                if source_kp:
                    result.append({
                        'relation_id': rel.id,
                        'source_kp_id': source_kp.id,
                        'source_kp_question': source_kp.question,
                        'source_chapter_id': source_kp.chapter_id,
                        'source_chapter_name': source_kp.chapter.name if source_kp.chapter else '',
                        'source_item_id': source_item.id,
                        'source_item_content': source_item.content
                    })
    
    return jsonify({'status': 'ok', 'data': result})

# 删除记忆记录（移出记忆规划）
@app.route('/api/memory-record/<int:kp_id>', methods=['DELETE'])
def delete_memory_record(kp_id):
    """从记忆规划中删除指定知识点"""
    user_id = request.args.get('user_id', 'default_user')
    
    # 查找记忆记录
    record = MemoryRecord.query.filter_by(user_id=user_id, knowledge_point_id=kp_id).first()
    if record:
        db.session.delete(record)
        db.session.commit()
        return jsonify({'status': 'ok', 'message': '记忆记录已删除'})
    else:
        return jsonify({'status': 'error', 'message': '未找到记忆记录'}), 404

# 获取记忆记录
@app.route('/api/memory-records', methods=['GET', 'POST'])
def handle_memory_records():
    user_id = request.args.get('user_id', 'default_user')
    
    if request.method == 'GET':
        kp_id = request.args.get('kp_id', type=int)
        status = request.args.get('status')
        
        query = MemoryRecord.query.filter_by(user_id=user_id)
        if kp_id:
            query = query.filter_by(knowledge_point_id=kp_id)
        if status:
            query = query.filter_by(status=status)
        
        records = query.all()
        return jsonify({'status': 'ok', 'data': [r.to_dict() for r in records]})
    else:
        data = request.json
        kp_ids = data.get('kp_ids', [])
        
        for kp_id in kp_ids:
            existing = MemoryRecord.query.filter_by(user_id=user_id, knowledge_point_id=kp_id).first()
            if not existing:
                record = MemoryRecord(
                    user_id=user_id,
                    knowledge_point_id=kp_id,
                    status='learning',
                    consecutive_correct=0,
                    interval_days=1,
                    next_review_date=datetime.now().date(),
                    learning_repetition=0,  # 新添加的字段
                    today_consecutive_count=0  # 新添加的字段
                )
                db.session.add(record)
        
        db.session.commit()
        return jsonify({'status': 'ok', 'message': 'Memory records created'})

# 获取今日待复习任务
@app.route('/api/memory/today-tasks', methods=['GET'])
def get_today_tasks():
    user_id = request.args.get('user_id', 'default_user')
    today = datetime.now().date()
    
    # 获取今日待复习的知识点
    # 优先显示所有初学中的记录（初学中都是当天需要完成的）
    learning_records = MemoryRecord.query.filter(
        MemoryRecord.user_id == user_id,
        MemoryRecord.status == 'learning'
    ).order_by(MemoryRecord.last_review_date.is_(None), MemoryRecord.last_review_date).all()
    
    reviewing_records = MemoryRecord.query.filter(
        MemoryRecord.user_id == user_id,
        MemoryRecord.status == 'reviewing',
        MemoryRecord.next_review_date <= today
    ).order_by(MemoryRecord.next_review_date).all()
    
    # 合并并优先学习中
    records = learning_records + reviewing_records
    
    # 获取对应的知识点详情
    kp_ids = [r.knowledge_point_id for r in records]
    kps = KnowledgePoint.query.filter(KnowledgePoint.id.in_(kp_ids)).all()
    kp_dict = {kp.id: kp.to_dict() for kp in kps}
    
    result = []
    for record in records:
        kp = kp_dict.get(record.knowledge_point_id)
        if kp:
            item = record.to_dict()
            item['knowledge_point'] = kp
            result.append(item)
    
    return jsonify({'status': 'ok', 'data': result})

# 记忆反馈（艾宾浩斯曲线核心逻辑）
@app.route('/api/memory/feedback', methods=['POST'])
def memory_feedback():
    """
    记忆反馈算法（完全按照需求实现）：
    
    一、初学中状态：
       - "背出了"：today_consecutive_count + 1，安排在4个题目之后再次出现
         → 连续2次"背出了"后，进入复习状态，下次复习为1天后
       - "模糊"/"背不出"：today_consecutive_count清零，安排在3个题目之后再次出现
         → 继续循环，直到达成连续2次"背出了"
    
    二、复习中状态：
       - "背出了"：consecutive_correct + 1，间隔前进到下一档位
         → 档位：1→2→4→7→15→30→60→120...天
         → 若已达最大档位，按1.5倍缓慢增加，最大365天
       - "模糊"：consecutive_correct回退1档（最少为1），间隔相应缩短
       - "背不出"：打回初学状态，重新开始短间隔循环
    
    三、下次复习日 = 当前日期 + 新间隔
    """
    user_id = request.args.get('user_id', 'default_user')
    data = request.json
    kp_id = data.get('kp_id')
    feedback = data.get('feedback')  # 'remembered', 'fuzzy', 'forgot'
    
    record = MemoryRecord.query.filter_by(user_id=user_id, knowledge_point_id=kp_id).first()
    if not record:
        return jsonify({'status': 'error', 'message': 'Memory record not found'}), 404
    
    today = datetime.now().date()
    
    # 复习间隔档位：1, 2, 4, 7, 15, 30, 60, 120, 180, 365
    intervals = [1, 2, 4, 7, 15, 30, 60, 120, 180, 365]
    
    if record.status == 'learning':
        # ==========================================
        # 【初学中状态】短间隔循环
        # ==========================================
        if feedback == 'remembered':
            # "背出了"：today_consecutive_count + 1
            record.today_consecutive_count += 1
            
            if record.today_consecutive_count >= 2:
                # ✓ 连续两次"背出了"，进入复习阶段
                record.status = 'reviewing'
                record.consecutive_correct = 1  # 复习阶段从1开始
                record.interval_days = 1  # 对应1天间隔
                record.next_review_date = today + timedelta(days=1)
                record.learning_repetition = 0
                record.today_consecutive_count = 0  # 重置连续计数
            else:
                # 继续初学循环，安排在4个题目之后
                record.learning_repetition = 4
                record.next_review_date = today  # 今天内循环
        else:
            # "模糊"或"背不出"：today_consecutive_count清零，安排在3个题目之后
            record.today_consecutive_count = 0
            record.learning_repetition = 3
            record.next_review_date = today  # 今天内循环
            
    else:
        # ==========================================
        # 【复习中状态】间隔复习
        # ==========================================
        if feedback == 'remembered':
            # "背出了"：consecutive_correct + 1，间隔前进
            # 同一天只能增加一个档位，同一天多次选择背出了也只增加一档
            if record.last_review_date == today:
                # 今天已经复习过了，今天已经增加过档位了，不再重复增加
                pass
            else:
                # 今天第一次背出了，增加一档
                record.consecutive_correct += 1
            
            # 计算新间隔（consecutive_correct从1开始，所以索引是consecutive_correct-1）
            if record.consecutive_correct <= len(intervals):
                # 在标准档位范围内
                record.interval_days = intervals[record.consecutive_correct - 1]
            else:
                # 已达最大档位，按1.5倍缓慢增加，最大365天
                record.interval_days = min(int(record.interval_days * 1.5), 365)
            
            record.next_review_date = today + timedelta(days=record.interval_days)
            
        elif feedback == 'fuzzy':
            # "模糊"：回退一档（最少保留为1）
            record.consecutive_correct = max(1, record.consecutive_correct - 1)
            record.interval_days = intervals[record.consecutive_correct - 1]
            record.next_review_date = today + timedelta(days=record.interval_days)
            
        else:
            # "背不出"：打回初学状态，重新开始短间隔循环
            record.status = 'learning'
            record.consecutive_correct = 0
            record.interval_days = 1
            record.next_review_date = today
            record.learning_repetition = 3  # 安排在3个题目之后出现
            record.today_consecutive_count = 0  # 重置连续计数
    
    record.last_review_date = today
    record.review_count += 1
    
    db.session.commit()
    return jsonify({'status': 'ok', 'data': record.to_dict()})

# 更新学习循环进度（练习时调用，每练习一个题目减1）
@app.route('/api/memory/advance-repetition', methods=['POST'])
def advance_repetition():
    """减少一个知识点的learning_repetition计数"""
    user_id = request.args.get('user_id', 'default_user')
    data = request.json
    kp_id = data.get('kp_id')
    
    record = MemoryRecord.query.filter_by(user_id=user_id, knowledge_point_id=kp_id).first()
    if not record:
        return jsonify({'status': 'error', 'message': 'Memory record not found'}), 404
    
    if record.learning_repetition > 0:
        record.learning_repetition -= 1
        db.session.commit()
    
    return jsonify({'status': 'ok', 'data': record.to_dict()})

# 获取学习统计
@app.route('/api/memory/statistics', methods=['GET'])
def get_memory_statistics():
    user_id = request.args.get('user_id', 'default_user')
    
    total_records = MemoryRecord.query.filter_by(user_id=user_id).count()
    learning_count = MemoryRecord.query.filter_by(user_id=user_id, status='learning').count()
    reviewing_count = MemoryRecord.query.filter_by(user_id=user_id, status='reviewing').count()
    
    today = datetime.now().date()
    
    # 今日待复习：初学中所有记录 + 复习中逾期未做的记录
    # 初学中状态：只要是learning状态，当天都需要复习
    learning_today = MemoryRecord.query.filter_by(
        user_id=user_id,
        status='learning'
    ).count()
    
    # 复习中逾期
    reviewing_overdue = MemoryRecord.query.filter(
        MemoryRecord.user_id == user_id,
        MemoryRecord.status == 'reviewing',
        MemoryRecord.next_review_date <= today
    ).count()
    
    today_review_count = learning_today + reviewing_overdue
    
    return jsonify({
        'status': 'ok',
        'data': {
            'total_records': total_records,
            'learning_count': learning_count,
            'reviewing_count': reviewing_count,
            'today_review_count': today_review_count
        }
    })

# 清空全部记忆规划
@app.route('/api/memory/clear-all', methods=['DELETE'])
def clear_all_memory_records():
    try:
        user_id = request.args.get('user_id', 'default_user')
        MemoryRecord.query.filter_by(user_id=user_id).delete()
        db.session.commit()
        return jsonify({'status': 'ok', 'message': '清空成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# 获取记忆规划预览数据（用于首页和数据统计）
@app.route('/api/memory/preview', methods=['GET'])
def get_memory_preview():
    user_id = request.args.get('user_id', 'default_user')
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)

    # 1. 知识点已加入规划总条数
    total_records = MemoryRecord.query.filter_by(user_id=user_id).count()

    # 2. 今日预计复习条数（初学中全部 + 复习中今天到期）
    learning_today = MemoryRecord.query.filter_by(
        user_id=user_id, status='learning'
    ).count()
    reviewing_today = MemoryRecord.query.filter(
        MemoryRecord.user_id == user_id,
        MemoryRecord.status == 'reviewing',
        MemoryRecord.next_review_date <= today
    ).count()
    today_review_count = learning_today + reviewing_today

    # 3. 明日预计复习条数
    learning_tomorrow = MemoryRecord.query.filter_by(
        user_id=user_id, status='learning'
    ).count()
    reviewing_tomorrow = MemoryRecord.query.filter(
        MemoryRecord.user_id == user_id,
        MemoryRecord.status == 'reviewing',
        MemoryRecord.next_review_date == tomorrow
    ).count()
    tomorrow_review_count = learning_tomorrow + reviewing_tomorrow

    # 4. 近15天需要复习的条数（按天分组）
    next_15_days = []
    for i in range(1, 16):
        day = today + timedelta(days=i)
        learning_count = MemoryRecord.query.filter_by(
            user_id=user_id, status='learning'
        ).count()
        reviewing_count = MemoryRecord.query.filter(
            MemoryRecord.user_id == user_id,
            MemoryRecord.status == 'reviewing',
            MemoryRecord.next_review_date == day
        ).count()
        next_15_days.append({
            'date': day.strftime('%m-%d'),
            'count': learning_count + reviewing_count
        })

    return jsonify({
        'status': 'ok',
        'data': {
            'total_records': total_records,
            'today_review_count': today_review_count,
            'tomorrow_review_count': tomorrow_review_count,
            'next_15_days': next_15_days
        }
    })


# 获取今日需要复习的知识点（用于默写练习）
@app.route('/api/memory/today-review', methods=['GET'])
def get_today_review_for_dictation():
    user_id = request.args.get('user_id', 'default_user')
    
    today = datetime.now().date()
    
    # 获取所有需要复习的记录
    # 1. 初学中的记录（learning状态）- 全部返回
    learning_records = MemoryRecord.query.filter_by(
        user_id=user_id,
        status='learning'
    ).all()
    
    # 2. 复习中已到期的记录（下次复习日期 <= 今天）
    reviewing_records = MemoryRecord.query.filter(
        MemoryRecord.user_id == user_id,
        MemoryRecord.status == 'reviewing',
        MemoryRecord.next_review_date <= today
    ).all()
    
    # 3. 今天已完成复习的记录（上次复习日期 == 今天）
    completed_today_records = MemoryRecord.query.filter(
        MemoryRecord.user_id == user_id,
        MemoryRecord.last_review_date == today
    ).all()
    
    # 合并并获取知识点详情（去重）
    kp_ids = []
    seen_ids = set()
    for r in learning_records + reviewing_records + completed_today_records:
        if r.knowledge_point_id not in seen_ids:
            kp_ids.append(r.knowledge_point_id)
            seen_ids.add(r.knowledge_point_id)
    
    if not kp_ids:
        return jsonify({'status': 'ok', 'data': []})
    
    # 获取知识点详情（包含题目和答案）
    knowledge_points = KnowledgePoint.query.filter(
        KnowledgePoint.id.in_(kp_ids)
    ).all()
    
    result = []
    for kp in knowledge_points:
        kp_dict = kp.to_dict()
        # 获取条目详情
        items = KnowledgePointItem.query.filter_by(knowledge_point_id=kp.id).order_by(KnowledgePointItem.order).all()
        kp_dict['items'] = [item.to_dict() for item in items]
        result.append(kp_dict)
    
    return jsonify({'status': 'ok', 'data': result})

# 随机获取知识点（用于默写练习）
@app.route('/api/knowledge-points/random', methods=['GET'])
def get_random_knowledge_points():
    count = request.args.get('count', 10, type=int)
    chapter_ids = request.args.getlist('chapter_id', type=int)
    priority = request.args.get('priority')
    
    query = KnowledgePoint.query.filter_by(status='active')
    
    if chapter_ids:
        query = query.filter(KnowledgePoint.chapter_id.in_(chapter_ids))
    if priority:
        query = query.filter_by(priority=priority)
    
    all_kps = query.all()
    
    # 随机选择
    if len(all_kps) <= count:
        selected = all_kps
    else:
        selected = random.sample(all_kps, count)
    
    return jsonify({
        'status': 'ok',
        'data': [kp.to_dict() for kp in selected]
    })

# 最后添加：为Vue Router历史模式配置catch-all路由
@app.route('/<path:path>', methods=['GET'])
def catch_all(path):
    # 如果是API请求，让它继续到404
    if path.startswith('api/'):
        return jsonify({'status': 'error', 'message': 'Not found'}), 404
    
    # 处理静态文件（assets目录）
    if path.startswith('assets/'):
        file_path = os.path.join(app.static_folder, path)
        if os.path.exists(file_path) and os.path.isfile(file_path):
            return send_from_directory(app.static_folder, path)
    
    # 所有其他请求返回index.html，让Vue Router处理路由
    return send_from_directory(app.static_folder, 'index.html')


# 注册辨析判断模块 API
# =========== 辨析判断模块 API ===========
from flask import Blueprint, request, jsonify
from models import db, DistinguishQuestion, DistinguishOption, DistinguishRecord
from datetime import datetime, timedelta
import random

distinguish_bp = Blueprint("distinguish", __name__, url_prefix="/api/distinguish")

@distinguish_bp.route("/save", methods=["POST"])
def save_distinguish():
    """保存辨析题（从刷题勾选弹窗保存）"""
    user_id = request.args.get("user_id", "default_user")
    data = request.json
    question_id = data.get("question_id")
    options = data.get("options", [])
    if not question_id:
        return jsonify({"status": "error", "message": "question_id is required"}), 400
    dq = DistinguishQuestion(user_id=user_id, question_id=question_id)
    db.session.add(dq)
    db.session.flush()
    for opt in options:
        record = DistinguishOption(
            distinguish_question_id=dq.id,
            option_key=opt.get("key", ""),
            option_text=opt.get("text", ""),
            is_correct=opt.get("is_correct", True),
            corrected_text=opt.get("corrected_text", None)
        )
        db.session.add(record)
    db.session.commit()
    return jsonify({"status": "ok", "data": dq.to_dict()})

@distinguish_bp.route("/check/<int:question_id>", methods=["GET"])
def check_distinguish(question_id):
    """检查题目是否已加入辨析题"""
    user_id = request.args.get("user_id", "default_user")
    dq = DistinguishQuestion.query.filter_by(user_id=user_id, question_id=question_id).first()
    return jsonify({
        "status": "ok",
        "exists": dq is not None,
        "distinguish_id": dq.id if dq else None
    })

@distinguish_bp.route("/questions", methods=["GET"])
def list_distinguish():
    """获取所有辨析题（支持筛选）"""
    user_id = request.args.get("user_id", "default_user")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    chapter_id = request.args.get("chapter_id", type=int)
    in_plan = request.args.get("in_plan")
    
    query = DistinguishQuestion.query.filter_by(user_id=user_id)
    
    if chapter_id:
        query = query.join(DistinguishQuestion.question).filter(Question.chapter_id == chapter_id)
    
    if in_plan is not None:
        if in_plan.lower() == "true":
            plan_ids = [r.distinguish_question_id for r in DistinguishRecord.query.filter_by(user_id=user_id).all()]
            query = query.filter(DistinguishQuestion.id.in_(plan_ids))
        elif in_plan.lower() == "false":
            plan_ids = [r.distinguish_question_id for r in DistinguishRecord.query.filter_by(user_id=user_id).all()]
            query = query.filter(DistinguishQuestion.id.not_in(plan_ids))
    
    pagination = query.order_by(DistinguishQuestion.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return jsonify({"status": "ok", "data": [q.to_dict() for q in pagination.items], "total": pagination.total, "pages": pagination.pages, "page": page})

@distinguish_bp.route("/questions/<int:qid>", methods=["DELETE", "PUT"])
def handle_distinguish_single(qid):
    """处理单个辨析题的DELETE和PUT请求"""
    if request.method == "DELETE":
        dq = DistinguishQuestion.query.get(qid)
        if not dq:
            return jsonify({"status": "error", "message": "not found"}), 404
        db.session.delete(dq)
        db.session.commit()
        return jsonify({"status": "ok", "message": "delete ok"})
    elif request.method == "PUT":
        import traceback
        print(f"[DEBUG] PUT request: qid={qid}")
        print(f"[DEBUG] request.json: {request.json}")
        
        try:
            dq = DistinguishQuestion.query.get(qid)
            print(f"[DEBUG] DistinguishQuestion found: {dq is not None}")
            
            if not dq:
                return jsonify({"status": "error", "message": "not found"}), 404
            
            data = request.json
            if "stem" in data:
                dq.stem = data["stem"]
            if "explanation" in data:
                dq.explanation = data["explanation"]
            
            if "options" in data:
                DistinguishOption.query.filter_by(distinguish_question_id=qid).delete()
                print(f"[DEBUG] Deleted old options")
                
                for opt in data["options"]:
                    new_opt = DistinguishOption(
                        distinguish_question_id=qid,
                        option_key=opt.get("option_key", ""),
                        option_text=opt.get("option_text", ""),
                        is_correct=opt.get("is_correct", True),
                        corrected_text=opt.get("corrected_text", None)
                    )
                    db.session.add(new_opt)
                print(f"[DEBUG] Added {len(data['options'])} new options")
            
            db.session.commit()
            print(f"[DEBUG] Commit successful")
            
            return jsonify({"status": "ok", "data": dq.to_dict()})
        except Exception as e:
            print(f"[DEBUG] Exception: {str(e)}")
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

@distinguish_bp.route("/questions/batch", methods=["DELETE"])
def batch_delete_distinguish():
    """批量删除辨析题"""
    ids = request.json.get("ids", [])
    if not ids:
        return jsonify({"status": "error", "message": "ids required"}), 400
    DistinguishQuestion.query.filter(DistinguishQuestion.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({"status": "ok", "message": f"deleted {len(ids)}"})

@distinguish_bp.route("/plan/add", methods=["POST"])
def add_to_plan():
    """添加选项到辨析规划"""
    user_id = request.args.get("user_id", "default_user")
    option_ids = request.json.get("option_ids", [])
    if not option_ids:
        return jsonify({"status": "error", "message": "option_ids required"}), 400
    today = datetime.now().date()
    count = 0
    for opt_id in option_ids:
        if DistinguishRecord.query.filter_by(user_id=user_id, option_id=opt_id).first():
            continue
        opt = DistinguishOption.query.get(opt_id)
        if not opt:
            continue
        db.session.add(DistinguishRecord(user_id=user_id, option_id=opt_id, status="learning", next_review_date=today))
        count += 1
    db.session.commit()
    return jsonify({"status": "ok", "message": f"add {count} to plan"})

@distinguish_bp.route("/plan/statistics", methods=["GET"])
def plan_statistics():
    """获取辨析规划统计"""
    user_id = request.args.get("user_id", "default_user")
    today = datetime.now().date()
    total = DistinguishRecord.query.filter_by(user_id=user_id).count()
    learning = DistinguishRecord.query.filter_by(user_id=user_id, status="learning").count()
    reviewing = DistinguishRecord.query.filter_by(user_id=user_id, status="reviewing").count()
    overdue = DistinguishRecord.query.filter(DistinguishRecord.user_id == user_id, DistinguishRecord.status == "reviewing", DistinguishRecord.next_review_date <= today).count()
    return jsonify({"status": "ok", "data": {"total_records": total, "learning_count": learning, "reviewing_count": reviewing, "today_review_count": learning + overdue}})

@distinguish_bp.route("/plan/tasks", methods=["GET"])
def plan_tasks():
    """获取今日辨析任务（包含重复题目）"""
    user_id = request.args.get("user_id", "default_user")
    today = datetime.now().date()
    
    learning = DistinguishRecord.query.filter_by(user_id=user_id, status="learning").all()
    reviewing = DistinguishRecord.query.filter(DistinguishRecord.user_id == user_id, DistinguishRecord.status == "reviewing", DistinguishRecord.next_review_date <= today).all()
    
    result = generate_unique_queue(learning, reviewing)
    
    return jsonify({"status": "ok", "data": [t.to_dict() for t in result], "total": len(result)})


def generate_practice_queue(learning, reviewing, record_id=None, feedback=None):
    """生成练习队列（包含重复题目）"""
    queue = list(learning) + list(reviewing)
    
    if not queue:
        return []
    
    unique_ids = set()
    result = []
    
    for item in queue:
        if item.id not in unique_ids:
            result.append(item)
            unique_ids.add(item.id)
    
    if record_id and feedback:
        record = None
        for item in learning + reviewing:
            if item.id == record_id:
                record = item
                break
        
        if record and record.status == "learning":
            if feedback == "forgot":
                result.append(record)
            elif feedback == "remembered" and record.today_consecutive_count < 2:
                result.append(record)
    
    return result


def generate_unique_queue(learning, reviewing):
    """生成唯一题目队列（用于初始加载）"""
    queue = list(learning) + list(reviewing)
    
    if not queue:
        return []
    
    unique_ids = set()
    result = []
    
    for item in queue:
        if item.id not in unique_ids:
            result.append(item)
            unique_ids.add(item.id)
    
    return result

@distinguish_bp.route("/plan/feedback", methods=["POST"])
def plan_feedback():
    """辨析记忆反馈（艾宾浩斯遗忘曲线算法）"""
    user_id = request.args.get("user_id", "default_user")
    data = request.json
    record_id = data.get("record_id")
    fb = data.get("feedback")
    if not record_id or not fb:
        return jsonify({"status": "error", "message": "record_id and feedback required"}), 400
    record = DistinguishRecord.query.filter_by(id=record_id, user_id=user_id).first()
    if not record:
        return jsonify({"status": "error", "message": "not found"}), 404
    today = datetime.now().date()
    intervals = [1, 2, 4, 7, 15, 30, 60, 120, 180, 365]
    
    # 记录原始状态和连续正确次数（在状态改变之前）
    original_status = record.status
    original_consecutive_count = record.today_consecutive_count
    
    if record.status == "learning":
        if fb == "remembered":
            record.today_consecutive_count += 1
            
            if record.today_consecutive_count >= 2:
                record.status = "reviewing"
                record.consecutive_correct = 1
                record.interval_days = 1
                record.next_review_date = today + timedelta(days=1)
                record.today_consecutive_count = 0
        else:
            record.today_consecutive_count = 0
    else:
        if fb == "remembered":
            if record.last_review_date != today:
                record.consecutive_correct += 1
            if record.consecutive_correct <= len(intervals):
                record.interval_days = intervals[record.consecutive_correct - 1]
            else:
                record.interval_days = min(int(record.interval_days * 1.5), 365)
            record.next_review_date = today + timedelta(days=record.interval_days)
        else:
            record.status = "learning"
            record.consecutive_correct = 0
            record.interval_days = 1
            record.next_review_date = today
            record.today_consecutive_count = 0
    
    record.last_review_date = today
    record.review_count += 1
    db.session.commit()
    
    # 计算是否需要重复该题目以及间隔
    repeat_info = None
    if original_status == "learning":
        if fb == "forgot":
            repeat_info = {"interval": 8, "reason": "答错，8题后再次出现"}
        else:
            if original_consecutive_count == 0:
                repeat_info = {"interval": 12, "reason": "答对，12题后验证"}
    else:
        if fb == "forgot":
            repeat_info = {"interval": 8, "reason": "复习答错，打回初学，8题后再次出现"}

    return jsonify({
        "status": "ok", 
        "data": record.to_dict(),
        "repeat_info": repeat_info,
        "feedback_result": "remembered" if fb == "remembered" else "forgot"
    })

@distinguish_bp.route("/plan/practice", methods=["GET"])
def plan_practice():
    """获取辨析练习题目（包含重复题目）"""
    user_id = request.args.get("user_id", "default_user")
    today = datetime.now().date()
    
    learning = DistinguishRecord.query.filter_by(user_id=user_id, status="learning").all()
    reviewing = DistinguishRecord.query.filter(DistinguishRecord.user_id == user_id, DistinguishRecord.status == "reviewing", DistinguishRecord.next_review_date <= today).all()
    
    result = generate_practice_queue(learning, reviewing)
    
    return jsonify({"status": "ok", "data": [t.to_dict() for t in result], "total": len(result)})

@distinguish_bp.route("/plan/list", methods=["GET"])
def plan_list():
    """获取规划记录列表"""
    user_id = request.args.get("user_id", "default_user")
    records = DistinguishRecord.query.filter_by(user_id=user_id).order_by(DistinguishRecord.next_review_date).all()
    result = []
    for r in records:
        opt = DistinguishOption.query.get(r.option_id)
        if opt:
            dq = DistinguishQuestion.query.get(opt.distinguish_question_id)
            q = dq.question if dq else None
            result.append({
                "id": r.id,
                "option_id": r.option_id,
                "question_id": dq.id if dq else None,
                "chapter_name": q.chapter.name if q and q.chapter else "",
                "question_stem": q.stem if q else "",
                "option_key": opt.option_key,
                "option_text": opt.option_text,
                "is_correct": opt.is_correct,
                "status": r.status,
                "consecutive_correct": r.consecutive_correct,
                "interval_days": r.interval_days,
                "next_review_date": r.next_review_date.isoformat() if r.next_review_date else "",
                "today_consecutive_count": r.today_consecutive_count,
                "review_count": r.review_count
            })
    return jsonify({"status": "ok", "data": result, "total": len(result)})

@distinguish_bp.route("/plan/record/<int:rid>", methods=["DELETE"])
def delete_plan_record(rid):
    """从规划中移除"""
    record = DistinguishRecord.query.get(rid)
    if not record:
        return jsonify({"status": "error", "message": "not found"}), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify({"status": "ok", "message": "removed"})

@distinguish_bp.route("/plan/clear-all", methods=["DELETE"])
def clear_all_plan():
    """清除全部规划"""
    user_id = request.args.get("user_id", "default_user")
    DistinguishRecord.query.filter_by(user_id=user_id).delete()
    db.session.commit()
    return jsonify({"status": "ok", "message": "cleared"})

def register_distinguish_api(app):
    """注册辨析判断模块到Flask应用"""
    app.register_blueprint(distinguish_bp)


# =========== 刷题规划模块 API ===========

# 获取可加入规划的题目列表
@app.route('/api/practice-plan/questions', methods=['GET'])
def get_practice_plan_questions():
    """获取题库中的题目，用于选择加入规划"""
    user_id = request.args.get('user_id', 'default_user')
    chapter_ids = request.args.getlist('chapter_id', type=int)
    keyword = request.args.get('keyword', '')
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    
    query = Question.query.filter_by(status='published')
    
    if chapter_ids:
        query = query.filter(Question.chapter_id.in_(chapter_ids))
    
    if keyword:
        query = query.filter(Question.stem.like(f'%{keyword}%'))
    
    # 获取已加入规划的题目ID
    existing_question_ids = [r.question_id for r in PracticePlanRecord.query.filter_by(user_id=user_id).all()]
    
    total = query.count()
    questions = query.order_by(Question.id.asc()).offset((page - 1) * page_size).limit(page_size).all()
    
    result = []
    for q in questions:
        q_dict = q.to_dict()
        q_dict['already_in_plan'] = q.id in existing_question_ids
        result.append(q_dict)
    
    return jsonify({
        'status': 'ok',
        'data': result,
        'total': total,
        'page': page,
        'page_size': page_size
    })

# 添加题目到规划
@app.route('/api/practice-plan/add', methods=['POST'])
def add_to_practice_plan():
    """添加题目到刷题规划"""
    user_id = request.args.get('user_id', 'default_user')
    data = request.json
    question_ids = data.get('question_ids', [])
    
    if not question_ids:
        return jsonify({'status': 'error', 'message': 'question_ids required'}), 400
    
    today = datetime.now().date()
    added_count = 0
    skipped_count = 0
    
    for qid in question_ids:
        # 检查是否已存在
        existing = PracticePlanRecord.query.filter_by(user_id=user_id, question_id=qid).first()
        if existing:
            skipped_count += 1
            continue
        
        # 检查题目是否存在
        question = Question.query.get(qid)
        if not question:
            skipped_count += 1
            continue
        
        record = PracticePlanRecord(
            user_id=user_id,
            question_id=qid,
            status='learning',
            consecutive_correct=0,
            interval_days=1,
            next_review_date=today,
            last_review_date=None,
            review_count=0,
            learning_repetition=0,
            today_consecutive_count=0,
            correct_at_learning_count=0
        )
        db.session.add(record)
        added_count += 1
    
    db.session.commit()
    
    return jsonify({
        'status': 'ok',
        'message': f'添加成功 {added_count} 题，{skipped_count} 题已存在'
    })

# 获取规划记录列表
@app.route('/api/practice-plan/list', methods=['GET'])
def get_practice_plan_list():
    """获取刷题规划记录列表
    
    今日待复习任务：
    - 初学中状态：每天都需要复习
    - 复习中状态：next_review_date <= 今日
    """
    user_id = request.args.get('user_id', 'default_user')
    today = datetime.now().date()
    status_filter = request.args.get('status')
    
    # 获取初学中的题目（每天都需要复习，未完成的）
    query_learning = PracticePlanRecord.query.filter_by(user_id=user_id, status='learning', completed=False)
    
    # 获取复习中的题目（只在需要复习的日期出现，未完成的）
    query_reviewing = PracticePlanRecord.query.filter(
        PracticePlanRecord.user_id == user_id,
        PracticePlanRecord.status == 'reviewing',
        PracticePlanRecord.next_review_date <= today,
        PracticePlanRecord.completed == False
    )
    
    if status_filter == 'learning':
        records = query_learning.order_by(PracticePlanRecord.next_review_date).all()
    elif status_filter == 'reviewing':
        records = query_reviewing.order_by(PracticePlanRecord.next_review_date).all()
    else:
        # 合并两种状态的记录
        learning_records = query_learning.all()
        reviewing_records = query_reviewing.all()
        records = learning_records + reviewing_records
        records.sort(key=lambda x: x.next_review_date)
    
    return jsonify({
        'status': 'ok',
        'data': [r.to_dict() for r in records],
        'total': len(records)
    })

# 获取规划统计
@app.route('/api/practice-plan/statistics', methods=['GET'])
def get_practice_plan_statistics():
    """获取刷题规划统计"""
    user_id = request.args.get('user_id', 'default_user')
    today = datetime.now().date()
    
    # 总记录数：所有添加过的题目（不区分completed状态）
    total = PracticePlanRecord.query.filter_by(user_id=user_id).count()
    # 学习中的题目
    learning = PracticePlanRecord.query.filter_by(user_id=user_id, status='learning').count()
    # 复习中的题目
    reviewing = PracticePlanRecord.query.filter_by(user_id=user_id, status='reviewing').count()
    # 今日待复习：初学中 + 复习中且下次复习日期 <= 今日
    today_review = PracticePlanRecord.query.filter_by(user_id=user_id, status='learning').count()
    today_review += PracticePlanRecord.query.filter(
        PracticePlanRecord.user_id == user_id,
        PracticePlanRecord.status == 'reviewing',
        PracticePlanRecord.next_review_date <= today
    ).count()
    
    return jsonify({
        'status': 'ok',
        'data': {
            'total_records': total,
            'learning_count': learning,
            'reviewing_count': reviewing,
            'today_review_count': today_review
        }
    })

# 获取今日刷题任务（包含重复题目）
@app.route('/api/practice-plan/tasks', methods=['GET'])
def get_practice_plan_tasks():
    """获取今日刷题任务
    
    今日刷题任务：
    - 初学中状态：每天都需要复习
    - 复习中状态：next_review_date <= 今日
    """
    user_id = request.args.get('user_id', 'default_user')
    today = datetime.now().date()
    
    # 获取初学中状态的所有题目（未完成的）
    learning = PracticePlanRecord.query.filter_by(user_id=user_id, status='learning', completed=False).all()
    # 获取复习中且需要今日复习的题目
    reviewing = PracticePlanRecord.query.filter(
        PracticePlanRecord.user_id == user_id,
        PracticePlanRecord.status == 'reviewing',
        PracticePlanRecord.next_review_date <= today
    ).all()
    
    result = generate_practice_plan_queue(learning, reviewing)
    
    return jsonify({
        'status': 'ok',
        'data': [t.to_dict() for t in result],
        'total': len(result)
    })


def generate_practice_plan_queue(learning, reviewing):
    """生成刷题队列
    
    实现记忆算法中的间隔重复逻辑：
    - 初学中首次做对：安排在12题后再次出现
    - 初学中做错：安排在8题后再次出现
    - 复习中做错打回初学：安排在8题后再次出现
    """
    # 合并初学和复习中的题目
    queue = []
    seen_ids = set()
    
    for item in learning + reviewing:
        if item.id not in seen_ids:
            queue.append(item)
            seen_ids.add(item.id)
    
    # 根据 learning_repetition 字段调整题目位置
    items_to_reposition = []
    final_queue = []
    
    for item in queue:
        if item.learning_repetition and item.learning_repetition > 0:
            items_to_reposition.append(item)
        else:
            final_queue.append(item)
    
    # 将需要间隔出现的题目插入到指定位置
    for item in items_to_reposition:
        insert_pos = min(item.learning_repetition, len(final_queue))
        final_queue.insert(insert_pos, item)
    
    return final_queue


# 刷题反馈
@app.route('/api/practice-plan/feedback', methods=['POST'])
def practice_plan_feedback():
    """刷题记忆反馈
    
    初学中状态逻辑：
    - 首次做错：wrong_count++, 8题后再次出现
    - 首次做对：correct_at_learning_count=1, 12题后再次出现验证
    - 做对验证（第二次）：完成，移出任务列表
    - 做错验证：wrong_count++, 重置为8题后出现
    
    复习中状态逻辑：
    - 做对：完成，移出任务列表
    - 做错：打回初学中，重新开始
    """
    user_id = request.args.get('user_id', 'default_user')
    data = request.json
    record_id = data.get('record_id')
    feedback = data.get('feedback')  # 'correct' 或 'wrong'
    
    if not record_id or not feedback:
        return jsonify({'status': 'error', 'message': 'record_id and feedback required'}), 400
    
    record = PracticePlanRecord.query.filter_by(id=record_id, user_id=user_id).first()
    if not record:
        return jsonify({'status': 'error', 'message': 'not found'}), 404
    
    today = datetime.now().date()
    intervals = [1, 2, 4, 7, 15, 30, 60, 120, 180, 365]
    
    if record.status == 'learning':
        # ==========================================
        # 【初学中状态】短间隔循环
        # ==========================================
        if feedback == 'correct':
            record.correct_at_learning_count += 1
            
            if record.correct_at_learning_count >= 2:
                # ✓ 连续两次答对，完成初学，进入复习中状态
                record.status = 'reviewing'
                record.consecutive_correct = 1
                record.interval_days = 1
                record.next_review_date = today + timedelta(days=1)
                record.completed = True
                record.completed_at = datetime.now()
                record.learning_repetition = 0  # 重置，避免干扰后续队列排序
            else:
                # 第一次答对，安排在12个题目之后验证
                record.learning_repetition = 12
                record.next_review_date = today
        else:
            # 做错：重置验证计数，安排在8个题目之后出现
            record.correct_at_learning_count = 0
            record.learning_repetition = 8
            record.next_review_date = today
            
    else:
        # ==========================================
        # 【复习中状态】间隔复习
        # ==========================================
        if feedback == 'correct':
            # ✓ 复习题目做对一次，完成当日任务，增加间隔
            record.consecutive_correct += 1
            # 根据连续答对次数确定新间隔
            if record.consecutive_correct - 1 < len(intervals):
                record.interval_days = intervals[record.consecutive_correct - 1]
            else:
                # 超过最大档位，按1.5倍增加
                record.interval_days = int(record.interval_days * 1.5)
            # 设置下次复习日
            record.next_review_date = today + timedelta(days=record.interval_days)
            record.completed = True
            record.completed_at = datetime.now()
        else:
            # 做错：打回初学状态，重新开始短间隔循环
            record.status = 'learning'
            record.consecutive_correct = 0
            record.interval_days = 1
            record.next_review_date = today
            record.learning_repetition = 8  # 安排在8个题目之后出现
            record.today_consecutive_count = 0
            record.correct_at_learning_count = 0
            record.completed = False
    
    record.last_review_date = today
    record.review_count += 1
    
    db.session.commit()
    
    # 获取更新后的队列信息（只包含未完成的题目）
    learning = PracticePlanRecord.query.filter_by(user_id=user_id, status='learning', completed=False).all()
    reviewing = PracticePlanRecord.query.filter(
        PracticePlanRecord.user_id == user_id,
        PracticePlanRecord.status == 'reviewing',
        PracticePlanRecord.next_review_date <= today,
        PracticePlanRecord.completed == False
    ).all()
    
    remaining = len(learning) + len(reviewing)
    queue = generate_practice_plan_queue(learning, reviewing)
    
    return jsonify({
        'status': 'ok',
        'data': record.to_dict(),
        'remaining': remaining,
        'queue_length': len(queue)
    })

# 推进学习循环进度（每练习一个题目减1）
@app.route('/api/practice-plan/advance', methods=['POST'])
def advance_practice_repetition():
    """减少learning_repetition计数"""
    user_id = request.args.get('user_id', 'default_user')
    data = request.json
    record_id = data.get('record_id')
    
    record = PracticePlanRecord.query.filter_by(id=record_id, user_id=user_id).first()
    if not record:
        return jsonify({'status': 'error', 'message': 'not found'}), 404
    
    if record.learning_repetition > 0:
        record.learning_repetition -= 1
        db.session.commit()
    
    return jsonify({'status': 'ok', 'data': record.to_dict()})

# 从规划中移除
@app.route('/api/practice-plan/record/<int:rid>', methods=['DELETE'])
def delete_practice_plan_record(rid):
    """从刷题规划中移除"""
    record = PracticePlanRecord.query.get(rid)
    if not record:
        return jsonify({'status': 'error', 'message': 'not found'}), 404
    
    db.session.delete(record)
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'removed'})

# 清除全部规划
@app.route('/api/practice-plan/clear-all', methods=['DELETE'])
def clear_all_practice_plan():
    """清除全部刷题规划"""
    user_id = request.args.get('user_id', 'default_user')
    PracticePlanRecord.query.filter_by(user_id=user_id).delete()
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'cleared'})


# 注册辨析判断 Blueprint（只注册一次）
register_distinguish_api(app)

if __name__ == '__main__':
    print("=== 启动 Flask 服务器 ===")
    print(f"app.py 文件路径: {__file__}")
    print(f"调试模式: True")
    
    with app.app_context():
        db.create_all()
        if QuestionType.query.count() == 0:
            default_types = [
                QuestionType(name='单选题', is_multiple=False),
                QuestionType(name='多选题', is_multiple=True),
                QuestionType(name='判断题', is_multiple=False)
            ]
            db.session.add_all(default_types)
        if Chapter.query.count() == 0:
            default_chapter = Chapter(name='默认章节', order=0)
            db.session.add(default_chapter)
        if UserPreference.query.count() == 0:
            default_pref = UserPreference(user_id='default_user')
            db.session.add(default_pref)
        db.session.commit()

    app.run(debug=False, use_reloader=False, host='0.0.0.0', port=5001)
